"""수금현황 목록/검색 화면.

기간·건물·세입자별 수금 내역 조회 라우트(`/payments`)와 그 전용
도우미 함수들을 모아둔 모듈입니다. 수금 등록 화면은 `payment_register.py`,
건물/현세입자 조회 API는 `payments_api.py`에 있습니다.
"""
from calendar import monthrange
from datetime import date, datetime
import io

from flask import render_template, request, send_file

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import db
from app_instance import app
from utils import (
    buildings_and_rooms as _buildings_and_rooms,
    building_label as _building_label,
    clamp_date_str,
    first_date_for_tenant as _first_date_for_tenant,
    fmt_bunji_pair as _fmt_bunji_pair,
    fmt_date as _fmt_date,
    iso_min_date as _iso_min_date,
    login_required,
    lookup_current_tenant as _lookup_current_tenant,
    make_pager as _make_pager,
    pad_ipju_seq as _pad_ipju_seq,
    to_int_amt as _to_int_amt,
    paginate as _paginate,
    parse_bunji_src as _parse_bunji_src,
    tenant_key as _tenant_key,
)

# 인쇄 한 장(A4)에 담을 거래 줄 수 — 페이지 번호는 이 값으로 직접 나눠서 계산
# (브라우저 인쇄 엔진의 실제 쪽수 계산 CSS는 크롬이 지원 안 함)
_PRINT_ROWS_PER_PAGE = 35


def _ym_from_iso(s, today):
    try:
        df = date.fromisoformat((s or "")[:10])
        return str(df.year), f"{df.month:02d}"
    except ValueError:
        return str(today.year), f"{today.month:02d}"


def _tenant_status_sql(tenant_status, alias="d"):
    """현 거주자 / 과거 입주자 SQL 조건"""
    if tenant_status == "past":
        return f"({alias}.out_dt IS NOT NULL AND {alias}.out_dt >= '1000-01-01')"
    if tenant_status == "all":
        return "1=1"
    return f"({alias}.out_dt IS NULL OR {alias}.out_dt < '1000-01-01')"


def _tenant_is_current(row):
    """out_dt 없음·레거시 무효날짜 = 거주 중"""
    if not row:
        return False
    out = row.get("out_dt")
    if out is None:
        return True
    if isinstance(out, (datetime, date)):
        return out.year < 1000
    s = str(out).strip()
    return (not s) or s.startswith("0000") or s < "1000-01-01"


def _first_date_for_name(nm):
    """세입자 이름 최초 등장일: 입주일·수금일 중 가장 이른 날"""
    like = f"%{nm}%"
    row = db.query_one(
        """
        SELECT MIN(dt) AS mn FROM (
          SELECT d.ipju_dt AS dt
          FROM bd03_det d
          WHERE d.ipju_nm LIKE %s
            AND d.ipju_dt IS NOT NULL AND d.ipju_dt > '1000-01-01'
          UNION ALL
          SELECT s.sukum_dt AS dt
          FROM sukum01 s
          INNER JOIN bd03_det d
            ON d.bunji1=s.bunji1 AND d.bunji2=s.bunji2
           AND d.hosu_norm=s.hosu_norm AND d.ipju_seq=s.ipju_seq
          WHERE d.ipju_nm LIKE %s
            AND s.sukum_dt IS NOT NULL AND s.sukum_dt > '1000-01-01'
        ) t
        """,
        (like, like),
    )
    if not row or not row.get("mn"):
        return None
    return _iso_min_date(row["mn"])


def _payment_year_options(today):
    year_rows = db.query(
        """
        SELECT DISTINCT YEAR(sukum_dt) AS y
        FROM sukum01
        WHERE sukum_dt IS NOT NULL AND sukum_dt > '1000-01-01'
        ORDER BY y DESC
        """
    )
    years = [int(r["y"]) for r in year_rows if r.get("y")]
    if today.year not in years:
        years = [today.year] + years
    if today.year - 1 not in years:
        years.append(today.year - 1)
    return sorted(set(years), reverse=True)


def _empty_payment_filters(today):
    month_start = today.replace(day=1)
    return {
        "bunji1": "",
        "bunji2": "",
        "hosu": "",
        "ipju_seq": "",
        "name": "",
        "name_mode": "",
        "tenant_status": "current",
        "date_from": month_start.isoformat(),
        "date_to": today.isoformat(),
        "ym_year": str(today.year),
        "ym_month": f"{today.month:02d}",
        "all_hist": "",
        "include_dache": True,
    }


def _read_payment_list_args(args):
    bunji1, bunji2 = _parse_bunji_src(args)
    hosu = (args.get("hosu") or "").strip().upper()
    ipju_seq_f = _pad_ipju_seq(args.get("ipju_seq"))
    name_raw = (args.get("name") or args.get("q") or "").strip()
    name_mode = args.get("name_mode") == "1"
    tenant_status = (args.get("tenant_status") or "current").strip().lower()
    if tenant_status not in ("current", "past", "all"):
        tenant_status = "current"
    return {
        "bunji1": bunji1,
        "bunji2": bunji2,
        "hosu": hosu,
        "ipju_seq_f": ipju_seq_f,
        "name_raw": name_raw,
        "name_mode": name_mode,
        "name_q": name_raw if name_mode else "",
        "name_display": name_raw,
        "tenant_status": tenant_status,
        "date_from": (args.get("date_from") or "").strip(),
        "date_to": (args.get("date_to") or "").strip(),
        "ym_year": (args.get("ym_year") or "").strip(),
        "ym_month": (args.get("ym_month") or "").strip(),
        "use_ym": args.get("use_ym") == "1",
        "all_hist": args.get("all_hist") in ("1", "true", "yes"),
        "include_dache": (
            (args.getlist("include_dache")[-1] if args.getlist("include_dache") else "1")
            in ("1", "true", "yes", "on")
        ),
        "name_list_mode": bool(name_raw)
        and name_mode
        and not (bunji1 and bunji2 and hosu),
    }


def _resolve_list_period(
    *,
    today,
    use_ym,
    ym_year,
    ym_month,
    all_hist,
    name_list_mode,
    bunji1,
    bunji2,
    hosu,
    ipju_seq_f,
    name_q,
    date_from,
    date_to,
):
    month_start = today.replace(day=1)
    if use_ym and ym_year.isdigit() and ym_month.isdigit():
        y = int(ym_year)
        m = int(ym_month)
        if 1 <= m <= 12 and 1990 <= y <= 2100:
            last_day = monthrange(y, m)[1]
            date_from = date(y, m, 1).isoformat()
            month_end = date(y, m, last_day)
            date_to = min(month_end, today).isoformat()
            return date_from, date_to, str(y), f"{m:02d}"
        return date_from, date_to, ym_year, ym_month
    if all_hist and bunji1 and bunji2 and hosu:
        first = _first_date_for_tenant(bunji1, bunji2, hosu, ipju_seq_f)
        date_from = first or "2000-01-01"
        date_to = today.isoformat()
        ym_year, ym_month = _ym_from_iso(date_from, today)
        return date_from, date_to, ym_year, ym_month
    if name_list_mode:
        first = _first_date_for_name(name_q)
        date_from = first or "2000-01-01"
        date_to = today.isoformat()
        ym_year, ym_month = _ym_from_iso(date_from, today)
        return date_from, date_to, ym_year, ym_month
    if date_from:
        date_from = clamp_date_str(date_from)
    if date_to:
        date_to = clamp_date_str(date_to)
    if not date_from:
        if bunji1 and bunji2 and not hosu:
            # 건물 전체 수금 현황 바로가기: "이번달"만 보면 최근 납부가 없어
            # 보이는 경우가 많아서 최근 3개월로 넓게 잡음
            back = today.month - 3
            back_year = today.year
            while back < 1:
                back += 12
                back_year -= 1
            date_from = date(back_year, back, 1).isoformat()
        else:
            date_from = month_start.isoformat()
    if not date_to:
        date_to = today.isoformat()
    try:
        if date.fromisoformat(date_to[:10]) > today:
            date_to = today.isoformat()
    except ValueError:
        date_to = today.isoformat()
    ym_year, ym_month = _ym_from_iso(date_from, today)
    return date_from, date_to, ym_year, ym_month


def _search_tenants_by_name(name_q, tenant_status):
    """이름 매칭: 정확 일치 우선 → 접두 → 포함 (+ 현거주/과거 필터)"""
    status_sql = _tenant_status_sql(tenant_status, "d")
    tenants = db.query(
        f"""
        SELECT d.bunji1, d.bunji2, d.hosu, d.ipju_seq, d.ipju_nm,
               d.ipju_dt, d.out_dt, b.juso
        FROM bd03_det d
        LEFT JOIN bd01 b ON b.bunji1=d.bunji1 AND b.bunji2=d.bunji2
        WHERE TRIM(d.ipju_nm)=%s
          AND {status_sql}
        ORDER BY d.bunji1, d.bunji2, d.hosu, d.ipju_seq
        """,
        (name_q,),
    )
    if tenants:
        return tenants
    tenants = db.query(
        f"""
        SELECT d.bunji1, d.bunji2, d.hosu, d.ipju_seq, d.ipju_nm,
               d.ipju_dt, d.out_dt, b.juso
        FROM bd03_det d
        LEFT JOIN bd01 b ON b.bunji1=d.bunji1 AND b.bunji2=d.bunji2
        WHERE d.ipju_nm LIKE %s
          AND {status_sql}
        ORDER BY d.bunji1, d.bunji2, d.hosu, d.ipju_seq
        """,
        (f"{name_q}%",),
    )
    if tenants:
        return tenants
    return db.query(
        f"""
        SELECT d.bunji1, d.bunji2, d.hosu, d.ipju_seq, d.ipju_nm,
               d.ipju_dt, d.out_dt, b.juso
        FROM bd03_det d
        LEFT JOIN bd01 b ON b.bunji1=d.bunji1 AND b.bunji2=d.bunji2
        WHERE d.ipju_nm LIKE %s
          AND {status_sql}
        ORDER BY d.bunji1, d.bunji2, d.hosu, d.ipju_seq
        """,
        (f"%{name_q}%",),
    )


def _list_room_tenants(bunji1, bunji2, hosu, tenant_status):
    """한 호실의 입주 이력. 과거=퇴실자만, 전체=전원, 현=거주 중."""
    status_sql = _tenant_status_sql(tenant_status, "d")
    return db.query(
        f"""
        SELECT d.bunji1, d.bunji2, d.hosu, d.ipju_seq, d.ipju_nm,
               d.ipju_dt, d.out_dt, b.juso
        FROM bd03_det d
        LEFT JOIN bd01 b ON b.bunji1=d.bunji1 AND b.bunji2=d.bunji2
        WHERE d.bunji1=%s AND d.bunji2=%s
          AND d.hosu_norm=%s
          AND {status_sql}
        ORDER BY CAST(d.ipju_seq AS UNSIGNED) DESC
        """,
        (bunji1, bunji2, (hosu or "").strip().upper()),
    )


def _focus_single_tenant(t, name_q, today):
    bunji1 = t.get("bunji1") or ""
    bunji2 = t.get("bunji2") or ""
    hosu = (t.get("hosu") or "").strip().upper()
    ipju_seq_f = _pad_ipju_seq(t.get("ipju_seq"))
    name_display = (t.get("ipju_nm") or name_q).strip()
    first = _first_date_for_tenant(bunji1, bunji2, hosu, ipju_seq_f)
    date_from = first or "2000-01-01"
    date_to = today.isoformat()
    ym_year, ym_month = _ym_from_iso(date_from, today)
    return {
        "bunji1": bunji1,
        "bunji2": bunji2,
        "hosu": hosu,
        "ipju_seq_f": ipju_seq_f,
        "name_display": name_display,
        "date_from": date_from,
        "date_to": date_to,
        "ym_year": ym_year,
        "ym_month": ym_month,
        "all_hist": True,
    }


def _name_match_payment_groups(tenants, date_from, date_to):
    """동명이인: 수금 전건 로드 금지 + 통계 1회 쿼리로 묶기."""
    key_set = set()
    tenant_meta = []
    for t in tenants:
        t_b1 = t.get("bunji1") or ""
        t_b2 = t.get("bunji2") or ""
        t_hosu = (t.get("hosu") or "").strip().upper()
        t_seq = _pad_ipju_seq(t.get("ipju_seq"))
        key = (t_b1, t_b2, t_hosu, t_seq)
        if key in key_set:
            continue
        key_set.add(key)
        tenant_meta.append(
            {
                "bunji1": t_b1,
                "bunji2": t_b2,
                "hosu": t_hosu,
                "ipju_seq": t_seq,
                "ipju_nm": (t.get("ipju_nm") or "").strip(),
                "juso": (t.get("juso") or "").strip(),
                "ipju_dt": _iso_min_date(t.get("ipju_dt")),
                "out_dt": _iso_min_date(t.get("out_dt")),
                "is_current": _tenant_is_current(t),
            }
        )
    stats_map = {}
    if tenant_meta:
        placeholders = []
        args_in = []
        for m in tenant_meta:
            placeholders.append("(%s,%s,%s,%s)")
            args_in.extend([m["bunji1"], m["bunji2"], m["hosu"], m["ipju_seq"]])
        stats_sql = f"""
            SELECT
              s.bunji1, s.bunji2,
              s.hosu_norm AS hosu,
              s.ipju_seq AS ipju_seq,
              COUNT(*) AS hist_c,
              MIN(s.sukum_dt) AS mn,
              MAX(s.sukum_dt) AS mx,
              SUM(
                CASE
                  WHEN s.sukum_dt >= %s AND s.sukum_dt < %s + INTERVAL 1 DAY
                  THEN 1 ELSE 0
                END
              ) AS pay_c
            FROM sukum01 s
            WHERE (s.bunji1, s.bunji2, s.hosu_norm, s.ipju_seq)
                  IN ({','.join(placeholders)})
            GROUP BY s.bunji1, s.bunji2,
                     s.hosu_norm,
                     s.ipju_seq
        """
        stats_rows = db.query(
            stats_sql,
            tuple([date_from + " 00:00:00", date_to] + args_in),
        )
        for sr in stats_rows or []:
            stats_map[_tenant_key(sr.get("bunji1"), sr.get("bunji2"), sr.get("hosu"), sr.get("ipju_seq"))] = sr
    groups = []
    for m in tenant_meta:
        sr = stats_map.get((m["bunji1"], m["bunji2"], m["hosu"], m["ipju_seq"])) or {}
        hist_mn = sr.get("mn")
        hist_mx = sr.get("mx")
        groups.append(
            {
                "bunji1": m["bunji1"],
                "bunji2": m["bunji2"],
                "hosu": m["hosu"],
                "ipju_seq": m["ipju_seq"],
                "ipju_nm": m["ipju_nm"],
                "juso": m["juso"],
                "ipju_dt": m.get("ipju_dt"),
                "out_dt": m.get("out_dt"),
                "is_current": m.get("is_current"),
                "pay_count": int(sr.get("pay_c") or 0),
                "hist_count": int(sr.get("hist_c") or 0),
                "hist_from": str(hist_mn)[:10] if hist_mn is not None else None,
                "hist_to": str(hist_mx)[:10] if hist_mx is not None else None,
            }
        )
    return groups


def _payment_row_filters(
    bunji1, bunji2, hosu, ipju_seq_f, date_from, date_to,
    tenant_status="all",
    include_dache=False,
):
    where = ["1=1"]
    args = []
    if bunji1:
        where.append("s.bunji1=%s")
        args.append(bunji1)
    if bunji2:
        where.append("s.bunji2=%s")
        args.append(bunji2)
    if hosu:
        where.append("s.hosu_norm=%s")
        args.append(hosu)
    if ipju_seq_f:
        where.append("s.ipju_seq=%s")
        args.append(ipju_seq_f)
    elif tenant_status in ("current", "past"):
        # 순번 미지정: 현/과거 구분. 입주 이력 없는 수금은 제외
        where.append(_tenant_status_sql(tenant_status, "d"))
        where.append("d.ipju_seq IS NOT NULL")
    if date_from:
        where.append("s.sukum_dt >= %s")
        args.append(date_from + " 00:00:00")
    if date_to:
        where.append("s.sukum_dt < %s + INTERVAL 1 DAY")
        args.append(date_to)
    where.append("(s.del_yn IS NULL OR s.del_yn='' OR s.del_yn='N')")
    if not include_dache:
        # 실입만. 대체전표(종류 02 또는 실입 0·대체>0)는 「대체 포함」일 때만
        where.append(
            "NOT (s.sukum_gb='02' OR (COALESCE(s.su_dache_amt,0)>0 AND COALESCE(s.su_sil_amt,0)=0))"
        )
    return where, args


def _count_payment_rows(
    bunji1, bunji2, hosu, ipju_seq_f, date_from, date_to, all_hist,
    tenant_status="all",
    include_dache=False,
):
    where, args = _payment_row_filters(
        bunji1, bunji2, hosu, ipju_seq_f, date_from, date_to,
        tenant_status, include_dache,
    )
    need_tenant = (not ipju_seq_f) and tenant_status in ("current", "past")
    join_kw = "INNER JOIN" if need_tenant else "LEFT JOIN"
    row = db.query_one(
        f"""
        SELECT COUNT(*) AS c
        FROM sukum01 s
        {join_kw} bd03_det d
          ON d.bunji1=s.bunji1 AND d.bunji2=s.bunji2
         AND d.hosu_norm=s.hosu_norm AND d.ipju_seq=s.ipju_seq
        WHERE {' AND '.join(where)}
        """,
        tuple(args),
    )
    return int((row or {}).get("c") or 0)


def _query_payment_rows(
    bunji1, bunji2, hosu, ipju_seq_f, date_from, date_to, all_hist,
    tenant_status="all",
    include_dache=False,
    limit=None,
    offset=0,
):
    where, args = _payment_row_filters(
        bunji1, bunji2, hosu, ipju_seq_f, date_from, date_to,
        tenant_status, include_dache,
    )
    need_tenant = (not ipju_seq_f) and tenant_status in ("current", "past")
    join_kw = "INNER JOIN" if need_tenant else "LEFT JOIN"
    sql = f"""
        SELECT s.sukum_dt, s.sukum_seq, s.bunji1, s.bunji2, s.hosu, s.ipju_seq,
               s.sukum_char, s.sukum_gb, s.manage_desc,
               s.su_sil_amt, s.su_dache_amt, s.s_method, s.del_yn,
               c1.g_cd_nm AS char_nm, c2.g_cd_nm AS gb_nm,
               d.ipju_nm, d.ipju_dt, d.rent_amt, d.manage_amt, d.bojung_amt, d.yechi_amt, b.juso
        FROM sukum01 s
        {join_kw} bd03_det d
          ON d.bunji1=s.bunji1 AND d.bunji2=s.bunji2
         AND d.hosu_norm=s.hosu_norm AND d.ipju_seq=s.ipju_seq
        LEFT JOIN gicho_code c1
          ON c1.g_cd='01' AND c1.g_sub_cd=s.sukum_char
        LEFT JOIN gicho_code c2
          ON c2.g_cd='02' AND c2.g_sub_cd=s.sukum_gb
        LEFT JOIN bd01 b
          ON b.bunji1=s.bunji1 AND b.bunji2=s.bunji2
        WHERE {' AND '.join(where)}
        ORDER BY s.sukum_dt ASC, s.sukum_seq ASC
    """
    if limit is not None:
        sql += " LIMIT %s OFFSET %s"
        args = list(args) + [int(limit), int(offset or 0)]
    return db.query(sql, tuple(args))


@app.route("/payments")
@login_required
def payments():
    today = date.today()
    # 상단 메뉴 등: fresh=1 또는 쿼리 없음 → 이전 세입자/주소 조건 초기화
    arg_keys = [k for k in request.args.keys() if k not in ("fresh", "partial")]
    is_fresh = request.args.get("fresh") == "1" or len(arg_keys) == 0
    is_partial = (request.args.get("partial") or "").strip() in ("list", "result")
    if is_partial:
        buildings, rooms, years = [], [], []
    else:
        buildings, rooms = _buildings_and_rooms()
        years = _payment_year_options(today)

    if is_fresh:
        fresh_ctx = dict(
            payments=[],
            payment_groups=[],
            name_list_mode=False,
            picker_kind="",
            empty_hint="",
            is_fresh=True,
            buildings=buildings,
            rooms=rooms,
            years=years,
            filters=_empty_payment_filters(today),
            pager=None,
        )
        if (request.args.get("partial") or "").strip() in ("list", "result"):
            return render_template("payments_result.html", **fresh_ctx)
        return render_template("payments.html", **fresh_ctx)

    q = _read_payment_list_args(request.args)
    bunji1 = q["bunji1"]
    bunji2 = q["bunji2"]
    hosu = q["hosu"]
    ipju_seq_f = q["ipju_seq_f"]
    name_mode = q["name_mode"]
    name_q = q["name_q"]
    name_display = q["name_display"]
    tenant_status = q["tenant_status"]
    all_hist = q["all_hist"]
    include_dache = q["include_dache"]
    name_list_mode = q["name_list_mode"]

    date_from, date_to, ym_year, ym_month = _resolve_list_period(
        today=today,
        use_ym=q["use_ym"],
        ym_year=q["ym_year"],
        ym_month=q["ym_month"],
        all_hist=all_hist,
        name_list_mode=name_list_mode,
        bunji1=bunji1,
        bunji2=bunji2,
        hosu=hosu,
        ipju_seq_f=ipju_seq_f,
        name_q=name_q,
        date_from=q["date_from"],
        date_to=q["date_to"],
    )

    rows = []
    payment_groups = []
    picker_kind = "name" if name_list_mode else ""
    empty_hint = ""
    if name_list_mode:
        tenants = _search_tenants_by_name(name_q, tenant_status)
        if len(tenants) == 1:
            hit = _focus_single_tenant(tenants[0], name_q, today)
            bunji1 = hit["bunji1"]
            bunji2 = hit["bunji2"]
            hosu = hit["hosu"]
            ipju_seq_f = hit["ipju_seq_f"]
            name_display = hit["name_display"]
            date_from = hit["date_from"]
            date_to = hit["date_to"]
            ym_year = hit["ym_year"]
            ym_month = hit["ym_month"]
            all_hist = True
            name_list_mode = False
            picker_kind = ""
        elif len(tenants) >= 2:
            payment_groups = _name_match_payment_groups(tenants, date_from, date_to)

    # 주소+호실만 있고 순번이 없을 때: 현=현재 입주자, 과거/전체=그 호 입주자 목록
    if (
        not name_list_mode
        and bunji1
        and bunji2
        and hosu
        and not ipju_seq_f
    ):
        if tenant_status == "current":
            cur = _lookup_current_tenant(bunji1, bunji2, hosu)
            if cur and _tenant_is_current(cur):
                hit = _focus_single_tenant(cur, name_display, today)
                bunji1 = hit["bunji1"] or bunji1
                bunji2 = hit["bunji2"] or bunji2
                hosu = hit["hosu"] or hosu
                ipju_seq_f = hit["ipju_seq_f"]
                name_display = hit["name_display"]
                date_from = hit["date_from"]
                date_to = hit["date_to"]
                ym_year = hit["ym_year"]
                ym_month = hit["ym_month"]
                all_hist = True
            else:
                empty_hint = (
                    "현재 입주자가 없습니다. 「과거 입주자」에서 퇴실자를 고르세요."
                )
        else:
            tenants = _list_room_tenants(bunji1, bunji2, hosu, tenant_status)
            if len(tenants) == 1:
                hit = _focus_single_tenant(tenants[0], name_display, today)
                bunji1 = hit["bunji1"] or bunji1
                bunji2 = hit["bunji2"] or bunji2
                hosu = hit["hosu"] or hosu
                ipju_seq_f = hit["ipju_seq_f"]
                name_display = hit["name_display"]
                date_from = hit["date_from"]
                date_to = hit["date_to"]
                ym_year = hit["ym_year"]
                ym_month = hit["ym_month"]
                all_hist = True
            else:
                payment_groups = _name_match_payment_groups(
                    tenants, date_from, date_to
                )
                name_list_mode = True
                picker_kind = "room"

    pager = None
    if name_list_mode:
        payment_groups, pager = _paginate(payment_groups)
    elif not empty_hint:
        total = _count_payment_rows(
            bunji1,
            bunji2,
            hosu,
            ipju_seq_f,
            date_from,
            date_to,
            all_hist,
            tenant_status,
            include_dache,
        )
        pager = _make_pager(total)
        if total:
            rows = _query_payment_rows(
                bunji1,
                bunji2,
                hosu,
                ipju_seq_f,
                date_from,
                date_to,
                all_hist,
                tenant_status,
                include_dache,
                limit=pager["per_page"],
                offset=pager["offset"],
            )

    ctx = dict(
        payments=rows if not name_list_mode else [],
        payment_groups=payment_groups,
        name_list_mode=name_list_mode,
        picker_kind=picker_kind,
        empty_hint=empty_hint,
        is_fresh=False,
        buildings=buildings,
        rooms=rooms,
        years=years,
        pager=pager,
        filters={
            "bunji1": bunji1,
            "bunji2": bunji2,
            "hosu": hosu,
            "ipju_seq": ipju_seq_f,
            "name": name_display,
            "name_mode": "1" if name_mode else "",
            "tenant_status": tenant_status,
            "date_from": date_from,
            "date_to": date_to,
            "ym_year": ym_year,
            "ym_month": ym_month,
            "all_hist": "1" if all_hist else "",
            "include_dache": "1" if include_dache else "",
        },
    )
    if (request.args.get("partial") or "").strip() in ("list", "result"):
        return render_template("payments_result.html", **ctx)

    return render_template("payments.html", **ctx)


def _resolve_payment_print_context(args):
    """인쇄 미리보기·엑셀 다운로드가 공용으로 쓰는 조회 로직.

    /payments 화면과 동일한 필터(주소·호실·이름·기간 등)를 그대로 받아
    거래 내역과 합계를 계산해서 돌려준다. 이 함수 하나만 고치면
    인쇄와 엑셀 양쪽 결과가 항상 같이 맞는다.
    """
    today = date.today()
    q = _read_payment_list_args(args)
    bunji1 = q["bunji1"]
    bunji2 = q["bunji2"]
    hosu = q["hosu"]
    ipju_seq_f = q["ipju_seq_f"]
    tenant_status = q["tenant_status"]
    all_hist = q["all_hist"]
    include_dache = q["include_dache"]

    date_from, date_to, _ym_year, _ym_month = _resolve_list_period(
        today=today,
        use_ym=q["use_ym"],
        ym_year=q["ym_year"],
        ym_month=q["ym_month"],
        all_hist=all_hist,
        name_list_mode=False,
        bunji1=bunji1,
        bunji2=bunji2,
        hosu=hosu,
        ipju_seq_f=ipju_seq_f,
        name_q="",
        date_from=q["date_from"],
        date_to=q["date_to"],
    )
    rows = _query_payment_rows(
        bunji1, bunji2, hosu, ipju_seq_f, date_from, date_to, all_hist, tenant_status,
        include_dache,
    )

    total_sil = sum(int(r.get("su_sil_amt") or 0) for r in rows)
    total_dache = sum(int(r.get("su_dache_amt") or 0) for r in rows)
    total_deposit = 0
    total_rent = 0
    total_manage = 0
    for r in rows:
        char = str(r.get("sukum_char") or "").strip()
        if char in ("02", "03"):
            total_deposit += _to_int_amt(r.get("su_sil_amt")) + _to_int_amt(
                r.get("su_dache_amt")
            )
        elif char == "01":
            total_rent += _to_int_amt(r.get("rent_amt"))
            total_manage += _to_int_amt(r.get("manage_amt"))

    return {
        "rows": rows,
        "bunji1": bunji1,
        "bunji2": bunji2,
        "building_name": _building_label(bunji1, bunji2) if bunji1 and bunji2 else "",
        "addr_label": _fmt_bunji_pair(bunji1, bunji2) if bunji1 and bunji2 else "",
        "name_display": q["name_raw"],
        "date_from": date_from,
        "date_to": date_to,
        "total_sil": total_sil,
        "total_dache": total_dache,
        "total_deposit": total_deposit,
        "total_rent": total_rent,
        "total_manage": total_manage,
    }


@app.route("/payments/print")
@login_required
def payments_print():
    """수금(대체) 내역 인쇄 — 별도 인쇄 전용 템플릿.

    /payments 화면에서 지금 보고 있는 조건(주소·호실·이름·기간 등) 그대로
    받아서 거래 내역을 인쇄용으로 다시 조회함(화면과 같은 필터 로직 재사용).
    """
    ctx = _resolve_payment_print_context(request.args)
    rows = ctx["rows"]
    pages = [
        rows[i:i + _PRINT_ROWS_PER_PAGE]
        for i in range(0, len(rows), _PRINT_ROWS_PER_PAGE)
    ] or [[]]

    return render_template(
        "payments_print.html",
        building_name=ctx["building_name"],
        addr_label=ctx["addr_label"],
        name_display=ctx["name_display"],
        date_from=ctx["date_from"],
        date_to=ctx["date_to"],
        pages=pages,
        total_pages=len(pages),
        total_count=len(rows),
        total_sil=ctx["total_sil"],
        total_dache=ctx["total_dache"],
        total_deposit=ctx["total_deposit"],
        total_rent=ctx["total_rent"],
        total_manage=ctx["total_manage"],
    )


def _build_payments_excel(ctx):
    """수금 목록 rows → openpyxl 워크북. 인쇄용 화면과 같은 컬럼 구성."""
    wb = Workbook()
    ws = wb.active
    ws.title = "수금현황"

    thin_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    headers = [
        "수금일자", "번지", "호수", "성명", "입주일자",
        "실수금액", "대체금액", "보증/예치", "임대료", "관리비",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F3F4F6")
    center = Alignment(horizontal="center")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    money_cols = (6, 7, 8, 9, 10)  # 실수금액~관리비
    row_idx = 1
    for r in ctx["rows"]:
        row_idx += 1
        char = str(r.get("sukum_char") or "").strip()
        deposit = None
        rent = None
        manage = None
        if char in ("02", "03"):
            deposit = _to_int_amt(r.get("su_sil_amt")) + _to_int_amt(r.get("su_dache_amt"))
        elif char == "01":
            rent = _to_int_amt(r.get("rent_amt"))
            manage = _to_int_amt(r.get("manage_amt"))

        ws.append([
            _fmt_date(r.get("sukum_dt")),
            _fmt_bunji_pair(r.get("bunji1"), r.get("bunji2")),
            (r.get("hosu") or "").strip(),
            r.get("ipju_nm") or "",
            _fmt_date(r.get("ipju_dt")) if r.get("ipju_dt") else "",
            _to_int_amt(r.get("su_sil_amt")) or None,
            _to_int_amt(r.get("su_dache_amt")) or None,
            deposit,
            rent,
            manage,
        ])
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if col in money_cols:
                cell.number_format = "#,##0"

    # 합계 행 (엑셀 SUM 수식 — 원본 값이 바뀌어도 자동 재계산됨)
    total_row = row_idx + 1
    ws.cell(row=total_row, column=1, value=f"합 계 ({len(ctx['rows'])}건)").font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.border = thin_border
        if col in money_cols:
            col_letter = ws.cell(row=1, column=col).column_letter
            cell.value = f"=SUM({col_letter}2:{col_letter}{row_idx})" if row_idx >= 2 else 0
            cell.number_format = "#,##0"
        cell.font = Font(bold=True)

    widths = [12, 10, 8, 10, 12, 12, 12, 12, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{row_idx}"

    return wb


@app.route("/payments/print/excel")
@login_required
def payments_print_excel():
    """인쇄 미리보기와 동일한 조건의 수금 내역을 .xlsx로 다운로드."""
    ctx = _resolve_payment_print_context(request.args)
    wb = _build_payments_excel(ctx)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    label_bits = [b for b in (ctx["building_name"], ctx["name_display"]) if b]
    label = "_".join(label_bits) if label_bits else "전체"
    filename = f"수금현황_{label}_{ctx['date_from']}_{ctx['date_to']}.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
