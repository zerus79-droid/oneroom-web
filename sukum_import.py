"""수금 입금 자동반영 화면.

농협 등 은행 입출금 내역 파일(.xls/.xlsx)을 업로드하면 적요(입금자명)를
전체 건물의 현재 입주자 이름과 비교해 건물을 자동 감지하고, 그 건물의
입주자와 매칭한 뒤 선택한 건만 sukum01에 등록합니다.
"""
import io
import json
import os
import re
import time
import uuid
from datetime import date, datetime

from flask import flash, redirect, render_template, request, session, url_for

import db
from app_instance import app
from utils import (
    account_digits as _account_digits,
    building_label as _building_label,
    buildings_and_rooms as _buildings_and_rooms,
    calc_misu_amt as _calc_misu_amt,
    fmt_bunji_pair as _fmt_bunji_pair,
    login_required,
    make_pager as _make_pager,
    next_sukum_seq as _next_sukum_seq,
    pad_bunji as _pad_bunji,
    parse_page as _parse_page,
    require_write_access,
    table_columns as _table_columns,
)

try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None
from openpyxl import load_workbook

# 매칭 결과(입금 목록)를 새로고침해도 다시 안 나오게 세션 쿠키 대신
# 서버 임시 파일에 저장 (쿠키엔 담기엔 큼) — POST 응답을 바로 렌더하지 않고
# GET으로 리다이렉트(PRG 패턴)해서 새로고침·뒤로가기로 인한 재제출을 막음
_TMP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_import_tmp")
_TMP_MAX_AGE_SEC = 6 * 3600


def _tmp_path(token):
    safe = re.sub(r"[^a-f0-9]", "", token or "")
    return os.path.join(_TMP_DIR, f"{safe}.json") if safe else None


def _cleanup_tmp():
    try:
        now = time.time()
        for name in os.listdir(_TMP_DIR):
            p = os.path.join(_TMP_DIR, name)
            if now - os.path.getmtime(p) > _TMP_MAX_AGE_SEC:
                os.remove(p)
    except OSError:
        pass


def _save_state(bunji1, bunji2, deposits, auto_detected, filename="", account_no=""):
    os.makedirs(_TMP_DIR, exist_ok=True)
    _cleanup_tmp()
    token = uuid.uuid4().hex
    with open(_tmp_path(token), "w", encoding="utf-8") as f:
        json.dump(
            {
                "bunji1": bunji1,
                "bunji2": bunji2,
                "deposits": deposits,
                "auto_detected": auto_detected,
                "filename": filename or "",
                "account_no": account_no or "",
            },
            f,
            ensure_ascii=False,
        )
    return token


def _load_state(token):
    path = _tmp_path(token)
    if not path or not os.path.isfile(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return None


def _delete_state(token):
    path = _tmp_path(token)
    if path and os.path.isfile(path):
        try:
            os.remove(path)
        except OSError:
            pass


def _write_state(token, state):
    path = _tmp_path(token)
    if not path:
        return
    with open(path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False)


def _update_state_building(token, bunji1, bunji2):
    state = _load_state(token)
    if not state:
        return None
    state["bunji1"] = bunji1
    state["bunji2"] = bunji2
    state["auto_detected"] = False
    _write_state(token, state)
    return state


def _find_col(headers, *keys):
    for i, h in enumerate(headers):
        for k in keys:
            if k in h:
                return i
    return None


def _parse_amount(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def _parse_date_any(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = re.search(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})", str(v or ""))
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


_ACCOUNT_NO_RE = re.compile(r"\d[\d\-]{7,}\d")


def _extract_account_number(rows, header_row):
    """헤더 행 이전 부분(계좌정보 등)에서 계좌번호를 찾아 숫자만 반환.
    은행 파일은 보통 '계좌번호' 라벨과 실제 번호가 같은 행의 다른 칸에 들어있음
    (예: | 계좌번호 | | 356-1174-4206-13 | )."""
    scan_until = header_row if header_row is not None else min(15, len(rows))
    for row in rows[:scan_until]:
        if not row:
            continue
        cells = ["" if c is None else str(c).strip() for c in row[:10]]
        if not any("계좌" in c for c in cells):
            continue
        for c in cells:
            if "계좌" in c:
                continue
            m = _ACCOUNT_NO_RE.search(c)
            if m:
                return _account_digits(m.group())
    return None


def _extract_deposits_from_rows(rows, xlrd_book=None):
    header_row = None
    headers = []
    for i, row in enumerate(rows[:30]):
        vals = ["" if c is None else str(c).replace("\n", "").strip() for c in row[:14]]
        joined = " ".join(vals)
        if "거래일시" in joined and "입금" in joined:
            header_row = i
            headers = vals
            break
    account_no = _extract_account_number(rows, header_row)
    if header_row is None:
        return [], account_no

    col_dt = _find_col(headers, "거래일시")
    col_in = _find_col(headers, "입금금액", "입금액")
    col_name = _find_col(headers, "거래기록사항", "적요")
    if col_dt is None or col_in is None or col_name is None:
        return [], account_no

    deposits = []
    for row in rows[header_row + 1 :]:
        if col_dt >= len(row) or col_in >= len(row) or col_name >= len(row):
            continue
        raw_dt = row[col_dt]
        d = None
        if xlrd_book is not None and isinstance(raw_dt, float):
            try:
                t = xlrd.xldate_as_tuple(raw_dt, xlrd_book.datemode)
                d = date(t[0], t[1], t[2])
            except Exception:
                d = None
        if d is None:
            d = _parse_date_any(raw_dt)
        if not d:
            continue
        amount = _parse_amount(row[col_in])
        if amount <= 0:
            continue
        name = str(row[col_name] or "").strip()
        if not name or "예금이자" in name:
            continue
        deposits.append({"date": d.isoformat(), "amount": amount, "name": name})
    return deposits, account_no


def _patch_xlrd_object_errors():
    """농협 등 은행 .xls에 섞인 깨진 OBJECT 레코드(로고 등) 때문에 xlrd가
    'Unexpected data at end of OBJECT record'로 죽는 걸 막음 — 입금 데이터엔
    필요 없는 레코드라 통째로 무시."""
    if xlrd is None:
        return
    import xlrd.sheet as xlrd_sheet

    if getattr(xlrd_sheet.Sheet.handle_obj, "_patched", False):
        return

    def _ignore(self, data):
        return None

    _ignore._patched = True
    xlrd_sheet.Sheet.handle_obj = _ignore


def _load_bank_deposits_xls(raw_bytes):
    if xlrd is None:
        raise RuntimeError("xlrd 패키지가 설치되어 있지 않습니다.")
    _patch_xlrd_object_errors()
    book = xlrd.open_workbook(file_contents=raw_bytes, formatting_info=False)
    sheet = book.sheet_by_index(0)
    rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    return _extract_deposits_from_rows(rows, xlrd_book=book)


def _load_bank_deposits_xlsx(raw_bytes):
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    return _extract_deposits_from_rows(rows)


def load_bank_deposits(filename, raw_bytes):
    """반환: (deposits, account_no_digits_or_None)"""
    suffix = (filename or "").lower().rsplit(".", 1)[-1]
    if suffix == "xls":
        return _load_bank_deposits_xls(raw_bytes)
    if suffix in ("xlsx", "xlsm"):
        return _load_bank_deposits_xlsx(raw_bytes)
    raise ValueError("지원 형식: .xls / .xlsx")


def _name_core(s):
    t = (s or "").strip()
    t = re.sub(r"\(.*?\)", "", t)
    t = re.sub(r"\s+", "", t)
    return t


def _name_parts(s):
    """비교용 조각. 괄호 안(닫히지 않은 '김호진(행복요양' 포함)도 따로 둠."""
    t = re.sub(r"\s+", "", (s or "").strip())
    if not t:
        return []
    parts = [t]
    for m in re.finditer(r"\(([^)]+)\)", t):
        inner = (m.group(1) or "").strip()
        if len(inner) >= 2:
            parts.append(inner)
    m = re.search(r"\(([^)]+)$", t)
    if m:
        inner = (m.group(1) or "").strip()
        if len(inner) >= 2:
            parts.append(inner)
    outer = re.sub(r"\(.*", "", t).strip()
    if len(outer) >= 2:
        parts.append(outer)
    seen = []
    for p in parts:
        if p not in seen:
            seen.append(p)
    return seen


def _name_matches(deposit_name, tenant_name):
    ds = _name_parts(deposit_name)
    ts = _name_parts(tenant_name)
    for d in ds:
        for t in ts:
            if len(d) < 2 or len(t) < 2:
                continue
            if d == t or d in t or t in d:
                return True
    return False


def _rent_amount_matches(amount, rent, manage, tol_ratio=0.03, tol_min=1000):
    """engine.py의 correct_unit_by_rent_amount와 같은 방식 — 입금액이 그 호실
    월세+관리비랑 비슷하면(오차 3% 또는 1000원 중 큰 쪽) 그 호실로 신뢰."""
    total = int(rent or 0) + int(manage or 0)
    if total <= 0:
        return False
    tol = max(tol_min, total * tol_ratio)
    return abs(amount - total) <= tol


def _narrow_by_room_hint(text, candidates, amount=None, allow_bare_number=True):
    """적요 텍스트에 '401호'/'1층'/그냥 숫자 같은 호수 힌트가 있으면 후보를 그 1곳으로
    좁힌다. 못 좁히면 원래 후보 그대로 반환.
    예: '김호현(401호' → 401호, '김호현(1층임차' → 1층(첫자리 1인 호실),
    '가람606 703호월세관리비' → 703호, '가람501월세' → 501(숫자만 — 그 호실 월세+관리비가
    입금액이랑 비슷할 때만 신뢰, engine.py 방식과 동일)."""
    if len(candidates) <= 1:
        return candidates

    def hosu_of(c):
        return (c.get("hosu") or "").strip().upper()

    for m in re.finditer(r"(\d{2,4})\s*호", text):
        num = m.group(1).lstrip("0") or "0"
        hits = [c for c in candidates if hosu_of(c).lstrip("0") == num]
        if len(hits) == 1:
            return hits

    # '가람501월세' — 호수+월세는 강한 힌트. 금액이 달라도 그 호실로 확정
    # (금액 이상은 매칭 후 확인필요로 표시).
    for m in re.finditer(r"(\d{2,4})\s*월세", text):
        num = m.group(1).lstrip("0") or "0"
        hits = [c for c in candidates if hosu_of(c).lstrip("0") == num]
        if len(hits) == 1:
            return hits

    m = re.search(r"(\d)\s*층", text)
    if m:
        hits = [c for c in candidates if hosu_of(c)[:1] == m.group(1)]
        if len(hits) == 1:
            return hits

    if allow_bare_number:
        for m in re.finditer(r"\d{3}", text):
            hits = [c for c in candidates if hosu_of(c) == m.group(0)]
            if len(hits) == 1:
                c = hits[0]
                if amount is None or _rent_amount_matches(amount, c.get("rent_amt"), c.get("manage_amt")):
                    return hits

    return candidates


_EXCLUDE_SCOPE_READY = False


def _ensure_exclude_scope_cols():
    """제외 항목에 주소·계좌 범위를 붙인다. 컬럼이 있으면 ALTER 하지 않는다."""
    global _EXCLUDE_SCOPE_READY
    if _EXCLUDE_SCOPE_READY:
        return
    cols = _table_columns("sukum_import_exclude")
    if "bunji1" not in cols:
        db.execute(
            "ALTER TABLE sukum_import_exclude "
            "ADD COLUMN bunji1 CHAR(4) NOT NULL DEFAULT ''"
        )
    if "bunji2" not in cols:
        db.execute(
            "ALTER TABLE sukum_import_exclude "
            "ADD COLUMN bunji2 CHAR(4) NOT NULL DEFAULT ''"
        )
    if "acct_no" not in cols:
        db.execute(
            "ALTER TABLE sukum_import_exclude "
            "ADD COLUMN acct_no VARCHAR(32) NOT NULL DEFAULT ''"
        )
    _EXCLUDE_SCOPE_READY = True


def list_exclude_keywords():
    _ensure_exclude_scope_cols()
    rows = db.query(
        """
        SELECT e.id, e.keyword, e.bunji1, e.bunji2, e.acct_no, b.juso
        FROM sukum_import_exclude e
        LEFT JOIN bd01 b ON b.bunji1=e.bunji1 AND b.bunji2=e.bunji2
        ORDER BY e.keyword, e.bunji1, e.bunji2, e.id
        """
    ) or []
    for r in rows:
        r["bunji1"] = _pad_bunji(r.get("bunji1"))
        r["bunji2"] = _pad_bunji(r.get("bunji2")) if r.get("bunji1") else ""
        r["acct_no"] = _account_digits(r.get("acct_no") or "")
    return rows


def add_exclude_keyword(keyword, bunji1="", bunji2="", acct_no=""):
    _ensure_exclude_scope_cols()
    keyword = (keyword or "").strip()
    bunji1 = _pad_bunji(bunji1)
    bunji2 = _pad_bunji(bunji2) if bunji1 else ""
    acct_no = _account_digits(acct_no or "")
    if not keyword and not bunji1 and not acct_no:
        return
    db.execute(
        "INSERT INTO sukum_import_exclude "
        "(keyword, bunji1, bunji2, acct_no, sys_dt, uid) "
        "VALUES (%s, %s, %s, %s, NOW(), %s)",
        (keyword, bunji1 or "", bunji2 or "", acct_no, session.get("sabun") or ""),
    )


def update_exclude_keyword(keyword_id, keyword, bunji1="", bunji2="", acct_no=""):
    _ensure_exclude_scope_cols()
    try:
        keyword_id = int(keyword_id or 0)
    except (TypeError, ValueError):
        return
    if keyword_id <= 0:
        return
    keyword = (keyword or "").strip()
    bunji1 = _pad_bunji(bunji1)
    bunji2 = _pad_bunji(bunji2) if bunji1 else ""
    acct_no = _account_digits(acct_no or "")
    if not keyword and not bunji1 and not acct_no:
        return
    db.execute(
        "UPDATE sukum_import_exclude "
        "SET keyword=%s, bunji1=%s, bunji2=%s, acct_no=%s, sys_dt=NOW(), uid=%s "
        "WHERE id=%s",
        (keyword, bunji1 or "", bunji2 or "", acct_no, session.get("sabun") or "", keyword_id),
    )


def delete_exclude_keyword(keyword_id):
    db.execute("DELETE FROM sukum_import_exclude WHERE id=%s", (keyword_id,))


def _matches_excluded(name, bunji1, bunji2, account_no, rules):
    """적요 글자 + (있으면) 주소 + (있으면) 계좌. 비어 있는 조건은 전체."""
    name = name or ""
    bunji1 = _pad_bunji(bunji1)
    bunji2 = _pad_bunji(bunji2)
    acct = _account_digits(account_no or "")
    for r in rules or []:
        kw = (r.get("keyword") or "").strip()
        rb1 = _pad_bunji(r.get("bunji1"))
        rb2 = _pad_bunji(r.get("bunji2"))
        racct = _account_digits(r.get("acct_no") or "")
        if not kw and not rb1 and not racct:
            continue
        if kw and kw not in name:
            continue
        if rb1 and (rb1 != bunji1 or rb2 != bunji2):
            continue
        if racct and racct != acct:
            continue
        return True
    return False


def buildings_by_account(account_no):
    """계좌번호(숫자만)를 쓰는 건물 목록. 책임관리는 여러 건물이 같은(관리사무소) 통장을
    같이 쓰는 게 정상이라 0곳/1곳/여러 곳 다 나올 수 있음."""
    if not account_no:
        return []
    rows = db.query("SELECT bunji1, bunji2, bank_cd FROM bd01 WHERE bank_cd IS NOT NULL AND bank_cd<>''")
    return [
        (r["bunji1"], r["bunji2"])
        for r in rows
        if _account_digits(r.get("bank_cd")) == account_no
    ]


def detect_building(account_no):
    """계좌번호로 건물이 정확히 1곳 특정될 때만 자동 확정. 그 외(0곳/여러 곳)는
    이름으로 전체 건물을 추측하지 않고 그냥 포기 — 건물을 직접 선택하게 함."""
    candidates = buildings_by_account(account_no)
    if len(candidates) == 1:
        return candidates[0]
    return "", ""


_AMOUNT_FLAG_MULTIPLE = 2  # 월세+관리비 기준액의 이 배수를 넘으면 "확인필요"로 표시


_UTILITY_RE = re.compile(r"수도|전기|가스|공과금|공공요금|난방|온수|한전|열요금")
_LUMP_RE = re.compile(r"보증|계약|잔금")


def _is_utility_desc(name):
    """수도·전기·가스·공과금 적요는 월세가 아님. '월세'가 같이 있으면 임대료로 본다."""
    t = re.sub(r"\s+", "", name or "")
    if not t or "월세" in t:
        return False
    return bool(_UTILITY_RE.search(t))


def _is_lump_desc(name):
    """보증·계약·잔금은 월세가 섞여 있어도 자동 반영하지 않는다. 호실만 고르면 들어간다."""
    t = re.sub(r"\s+", "", name or "")
    return bool(t and _LUMP_RE.search(t))


def _monthly_due(t):
    return int(t.get("rent_amt") or 0) + int(t.get("manage_amt") or 0)


def _alloc_amounts(dues, amount):
    """호실별 월세 기준으로 입금액을 나눔. n개월분이면 각 호 n개월, 아니면
    floor(입금/월세합)개월 + 잔액은 월세 큰 호실. 합은 항상 입금액."""
    amount = int(amount or 0)
    n = len(dues)
    if n <= 0:
        return []
    if n == 1:
        return [amount]
    if any(int(d or 0) <= 0 for d in dues):
        base = amount // n
        alloc = [base] * n
        alloc[0] += amount - base * n
        return alloc
    dues = [int(d or 0) for d in dues]
    total = sum(dues)
    months = 1
    matched = False
    for m in range(1, 25):
        if _rent_amount_matches(amount, m * total, 0):
            months = m
            matched = True
            break
    if not matched:
        months = max(1, amount // total) if total > 0 else 1
    alloc = [d * months for d in dues]
    leftover = amount - sum(alloc)
    if leftover:
        biggest = max(range(n), key=lambda i: dues[i])
        alloc[biggest] += leftover
        if alloc[biggest] < 0:
            rest = amount
            alloc = [0] * n
            for i in range(n - 1):
                take = min(max(dues[i], 0), rest)
                alloc[i] = take
                rest -= take
            alloc[-1] = rest
    return alloc


def _alloc_by_caps(caps, amount):
    """앞 호실부터 한도(미수)만큼 채우고 마지막 호실이 잔액. 합은 항상 입금액."""
    amount = int(amount or 0)
    n = len(caps)
    if n <= 0:
        return []
    if n == 1:
        return [amount]
    alloc = [0] * n
    rest = amount
    for i in range(n - 1):
        take = min(max(int(caps[i] or 0), 0), rest)
        alloc[i] = take
        rest -= take
    alloc[-1] = rest
    return alloc


def _alloc_lump(candidates, amount):
    """복수 호실에 입금액을 나눔. n개월분이면 각 호 n개월, 아니면
    floor(입금/월세합)개월 + 잔액은 월세 큰 호실."""
    if not candidates:
        return []
    if len(candidates) == 1:
        return [(candidates[0], int(amount))]
    ordered = sorted(candidates, key=_hosu_sort_key)
    alloc = _alloc_amounts([_monthly_due(c) for c in ordered], amount)
    return list(zip(ordered, alloc))


def _split_parts_from_options(room_options, amount):
    """2호실 선택 시 호실별 금액. 미수가 있으면 미수부터, 없으면 월세 기준."""
    rooms = [
        o for o in (room_options or [])
        if "|" in (o.get("value") or "")
    ]
    if not rooms:
        return {}
    amount = int(amount or 0)
    misus = [int(o.get("misu") or 0) for o in rooms]
    dues = [int(o.get("due") or 0) for o in rooms]
    if any(m > 0 for m in misus):
        alloc = _alloc_by_caps(misus, amount)
    else:
        alloc = _alloc_amounts(dues, amount)
    return {
        rooms[i]["value"]: alloc[i]
        for i in range(len(rooms))
        if alloc[i] > 0
    }


def _try_split_lump(candidates, amount):
    """같은 입금자 복수 호실 + 입금액이 각 호 월세합(또는 n개월)과 같으면 호실별로 나눔.
    예: 한현승 2,550,000 → 201 2,000,000 + B01 550,000."""
    if len(candidates) < 2:
        return None
    dues = [_monthly_due(c) for c in candidates]
    if any(d <= 0 for d in dues):
        return None
    total = sum(dues)
    if not any(
        _rent_amount_matches(int(amount), n * total, 0) for n in range(1, 25)
    ):
        return None
    return _alloc_lump(candidates, amount)


def _finish_match_row(
    dep, match, amount, candidates, all_room_options,
    existing_set, bunji1, bunji2, needs_pick,
):
    hosu = (match.get("hosu") or "").strip().upper() if match else ""
    ipju_seq = str(match.get("ipju_seq") or "").zfill(2) if match else ""
    row = {
        "date": dep["date"],
        "amount": amount,
        "name": dep["name"],
        "hosu": hosu,
        "ipju_seq": ipju_seq,
        "tenant_nm": match.get("ipju_nm") if match else "",
        "amount_flag": False,
        "needs_pick": needs_pick,
        "room_options": (
            _room_options(
                candidates, with_combo=True,
                bunji1=bunji1, bunji2=bunji2,
                as_of=date.fromisoformat(dep["date"]),
            )
            if needs_pick
            else (all_room_options if not match else [])
        ),
    }
    if needs_pick:
        row["status"] = "matched"
    elif not match:
        row["status"] = "unmatched"
    elif (hosu, ipju_seq, row["date"], int(amount)) in existing_set:
        row["status"] = "duplicate"
    else:
        row["status"] = "matched"
        expected = _monthly_due(match)
        if expected > 0 and amount > expected * _AMOUNT_FLAG_MULTIPLE:
            misu = _calc_misu_amt(
                bunji1, bunji2, hosu, ipju_seq,
                match.get("rent_amt"), match.get("manage_amt"),
                match.get("ipju_dt"), as_of=date.fromisoformat(dep["date"]),
            )
            if amount > expected + misu:
                row["amount_flag"] = True
                row["room_options"] = all_room_options
    return row


def _match_deposits(deposits, bunji1, bunji2, account_no=""):
    if not (bunji1 and bunji2):
        return []
    tenants = db.query(
        """
        SELECT hosu, ipju_seq, ipju_nm, rent_amt, manage_amt, ipju_dt
        FROM bd03_det
        WHERE bunji1=%s AND bunji2=%s AND (out_dt IS NULL OR out_dt < '1000-01-01')
        """,
        (bunji1, bunji2),
    )
    existing = db.query(
        """
        SELECT hosu, ipju_seq, DATE(sukum_dt) AS d, su_sil_amt
        FROM sukum01
        WHERE bunji1=%s AND bunji2=%s
          AND (del_yn IS NULL OR del_yn='N' OR del_yn='')
        """,
        (bunji1, bunji2),
    )
    existing_set = {
        (
            (r.get("hosu") or "").strip().upper(),
            str(r.get("ipju_seq") or "").zfill(2),
            r["d"].isoformat() if r.get("d") else "",
            int(r.get("su_sil_amt") or 0),
        )
        for r in existing
    }
    exclude_rules = list_exclude_keywords()
    all_room_options = _room_options(tenants)

    results = []
    for dep in deposits:
        if _matches_excluded(dep["name"], bunji1, bunji2, account_no, exclude_rules):
            continue  # 제외 목록에 걸리면 매칭 결과에 아예 표시하지 않음

        # 수도·전기·가스 적요는 호수 힌트가 있어도 월세로 넣지 않음
        if _is_utility_desc(dep["name"]):
            results.append(
                _finish_match_row(
                    dep, None, dep["amount"], tenants, all_room_options,
                    existing_set, bunji1, bunji2, needs_pick=False,
                )
            )
            continue

        name_candidates = [t for t in tenants if _name_matches(dep["name"], t.get("ipju_nm") or "")]
        if len(name_candidates) > 1:
            # 동명이인/복수호실 — 적요의 호수 힌트('401호', '1층', 숫자만)로 먼저 좁혀봄
            candidates = _narrow_by_room_hint(
                dep["name"], name_candidates, amount=dep["amount"], allow_bare_number=True
            )
            if len(candidates) > 1:
                one_amt = [
                    c for c in candidates
                    if _rent_amount_matches(dep["amount"], c.get("rent_amt"), c.get("manage_amt"))
                ]
                if len(one_amt) == 1:
                    candidates = one_amt
        elif not name_candidates:
            # 이름 매칭이 아예 없으면(적요가 '가람501월세'처럼 호수/건물 위주인 경우)
            # 건물 전체에서 호수 힌트로 좁혀봄. 숫자만 있는 힌트는 그 호실 월세+관리비가
            # 입금액이랑 비슷할 때만 신뢰(engine.py의 correct_unit_by_rent_amount 방식).
            candidates = _narrow_by_room_hint(
                dep["name"], tenants, amount=dep["amount"], allow_bare_number=True
            )
            if len(candidates) != 1:
                candidates = name_candidates
        else:
            candidates = name_candidates

        # 보증·계약·잔금(월세 섞인 경우 포함)은 자동 반영하지 않음 — 호실 선택 후 반영
        if _is_lump_desc(dep["name"]):
            match = candidates[0] if len(candidates) == 1 else None
            needs_pick = len(candidates) > 1
            row = _finish_match_row(
                dep, match, dep["amount"], candidates, all_room_options,
                existing_set, bunji1, bunji2, needs_pick=needs_pick,
            )
            if row["status"] == "matched" and not needs_pick:
                row["amount_flag"] = True
                row["room_options"] = all_room_options
            results.append(row)
            continue

        split = _try_split_lump(candidates, dep["amount"]) if len(candidates) > 1 else None
        if split:
            for match, amt in split:
                results.append(
                    _finish_match_row(
                        dep, match, amt, candidates, all_room_options,
                        existing_set, bunji1, bunji2, needs_pick=False,
                    )
                )
            continue

        match = candidates[0] if len(candidates) == 1 else None
        needs_pick = len(candidates) > 1
        results.append(
            _finish_match_row(
                dep, match, dep["amount"], candidates, all_room_options,
                existing_set, bunji1, bunji2, needs_pick=needs_pick,
            )
        )

    # 표시 순서: 미매칭 → 확인필요(금액 이상/복수호실) → 반영예정 → 날짜중복(이미 등록됨)
    def _sort_key(r):
        if r["status"] == "unmatched":
            return 0
        if r["status"] == "matched" and (r.get("amount_flag") or r.get("needs_pick")):
            return 1
        if r["status"] == "matched":
            return 2
        return 3  # duplicate

    results.sort(key=_sort_key)
    return results


def _hosu_sort_key(t):
    """지하(B) 먼저, 그다음 지상 1층→높은 층."""
    h = (t.get("hosu") if isinstance(t, dict) else t) or ""
    h = str(h).strip().upper()
    if h.startswith("B"):
        tail = h[1:].lstrip("0") or "0"
        try:
            return (0, int(tail), h)
        except ValueError:
            return (0, 0, h)
    digits = "".join(c for c in h if c.isdigit())
    try:
        return (1, int(digits or 0), h)
    except ValueError:
        return (2, 0, h)


def _room_options(tenants, with_combo=False, bunji1="", bunji2="", as_of=None):
    opts = []
    tenants = sorted(tenants, key=_hosu_sort_key)
    for t in tenants:
        h = (t.get("hosu") or "").strip().upper()
        seq = str(t.get("ipju_seq") or "").zfill(2)
        nm = (t.get("ipju_nm") or "").strip()
        if not h or not seq:
            continue
        misu = 0
        if bunji1 and bunji2 and as_of:
            misu = _calc_misu_amt(
                bunji1, bunji2, h, seq,
                t.get("rent_amt"), t.get("manage_amt"),
                t.get("ipju_dt"), as_of=as_of,
            )
        opts.append({
            "value": f"{h}|{seq}",
            "label": f"{h}호 {nm}".strip(),
            "misu": misu,
            "due": _monthly_due(t),
        })
    if with_combo and len(opts) >= 2:
        opts.append({"value": "ALL", "label": f"{len(opts)}호실"})
    return opts


def _insert_sukum(bunji1, bunji2, hosu, ipju_seq, sukum_dt, amount, name):
    hosu = (hosu or "").strip().upper()
    ipju_seq = str(ipju_seq or "").zfill(2)
    if not (hosu and ipju_seq) or int(amount or 0) <= 0:
        return False
    sukum_seq = _next_sukum_seq(sukum_dt, bunji1, bunji2, hosu)
    db.execute(
        """
        INSERT INTO sukum01 (
            sukum_dt, sukum_seq, bunji1, bunji2, hosu, ipju_seq,
            sukum_char, sukum_gb, manage_desc, su_sil_amt, su_dache_amt,
            suri_dt, suri_seq, s_method, del_yn, sys_dt, uid
        ) VALUES (
            %s, %s, %s, %s, %s, %s,
            '01', '03', %s, %s, 0,
            NULL, '', '', 'N', NOW(), %s
        )
        """,
        (
            sukum_dt + " 00:00:00",
            sukum_seq,
            bunji1,
            bunji2,
            hosu,
            ipju_seq,
            f"입금파일 자동반영 ({name or ''})",
            int(amount),
            session.get("sabun") or "",
        ),
    )
    return True


def _apply_selected(rows, bunji1, bunji2, selected_idx, manual_overrides, split_map=None):
    split_map = split_map or {}
    saved = 0
    applied = []
    for i, row in enumerate(rows):
        if row.get("status") == "duplicate":
            continue
        if i not in selected_idx:
            continue
        raw = manual_overrides.get(str(i)) or []
        if isinstance(raw, str):
            raw = [raw] if raw else []
        if any(v == "ALL" for v in raw):
            parts = dict(split_map.get(str(i)) or {})
            room_keys = [
                o.get("value")
                for o in (row.get("room_options") or [])
                if "|" in (o.get("value") or "")
            ]
            filled = {k: int(v) for k, v in parts.items() if int(v) > 0}
            if len(filled) == 1 and len(room_keys) >= 2:
                rest = int(row["amount"]) - next(iter(filled.values()))
                empty = [k for k in room_keys if k not in filled]
                if rest > 0 and len(empty) == 1:
                    parts[empty[0]] = rest
            if sum(int(v) for v in parts.values()) != int(row["amount"]):
                parts = _split_parts_from_options(
                    row.get("room_options"), row.get("amount")
                )
            if sum(int(v) for v in parts.values()) != int(row["amount"]):
                continue
            n = 0
            for key, amt in parts.items():
                if "|" not in key:
                    continue
                hosu, ipju_seq = key.split("|", 1)
                if _insert_sukum(
                    bunji1, bunji2, hosu, ipju_seq,
                    row["date"], amt, row.get("name"),
                ):
                    n += 1
                    saved += 1
            if n:
                applied.append(i)
            continue
        picks = []
        seen = set()
        for v in raw:
            if v and "|" in v and v not in seen:
                seen.add(v)
                picks.append(v)
        if len(picks) == 1:
            hosu, ipju_seq = picks[0].split("|", 1)
        else:
            hosu, ipju_seq = row.get("hosu"), row.get("ipju_seq")
        if _insert_sukum(
            bunji1, bunji2, hosu, ipju_seq,
            row["date"], row["amount"], row.get("name"),
        ):
            saved += 1
            applied.append(i)
    return saved, applied


@app.route("/payments/import", methods=["GET", "POST"])
@login_required
@require_write_access
def payments_import():
    """PRG(Post-Redirect-Get) 패턴: 업로드·매칭·반영은 전부 POST 후 GET으로
    리다이렉트한다 — 새로고침해도 파일이 재업로드/재반영되지 않도록 하기 위함.
    매칭 결과(입금 목록)는 쿠키 세션에 담기엔 커서 서버 임시 파일(_import_tmp/)에
    token으로 저장해두고 GET에서 그 token으로 불러와 다시 그린다."""
    buildings, _rooms = _buildings_and_rooms()

    if request.method == "POST" and request.form.get("action") == "apply":
        token = request.form.get("token") or ""
        state = _load_state(token)
        if not state:
            flash("매칭 결과가 만료됐습니다. 파일을 다시 올려주세요.", "err")
            return redirect(url_for("payments_import"))
        bunji1, bunji2 = state["bunji1"], state["bunji2"]
        selected_idx = set()
        one = (request.form.get("apply_one") or "").strip()
        if one != "":
            try:
                selected_idx.add(int(one))
            except ValueError:
                pass
        else:
            for v in request.form.getlist("apply_idx"):
                try:
                    selected_idx.add(int(v))
                except ValueError:
                    pass
        manual_overrides = {}
        for k in request.form:
            if not k.startswith("manual_"):
                continue
            vals = [v for v in request.form.getlist(k) if v]
            if vals:
                manual_overrides[k[len("manual_"):]] = vals
        split_map = {}
        for k in request.form:
            if not k.startswith("split_"):
                continue
            rest = k[len("split_"):]
            idx, sep, room = rest.partition("_")
            if not sep or "|" not in room:
                continue
            amt = _parse_amount(request.form.get(k))
            if amt > 0:
                split_map.setdefault(idx, {})[room] = amt
        rows = _match_deposits(
            state["deposits"], bunji1, bunji2, state.get("account_no") or "",
        )
        _saved, applied = _apply_selected(
            rows, bunji1, bunji2, selected_idx, manual_overrides, split_map,
        )
        if applied:
            drop = {
                (rows[i]["date"], rows[i]["name"], int(rows[i]["amount"]))
                for i in applied
            }
            state["deposits"] = [
                d for d in state["deposits"]
                if (d["date"], d["name"], int(d["amount"])) not in drop
            ]
            _write_state(token, state)
        if one != "" and not applied:
            return redirect(url_for("payments_import", token=token, err=one))
        return redirect(url_for("payments_import", token=token))

    if request.method == "POST" and request.form.get("action") == "parse":
        bunji1 = _pad_bunji(request.form.get("bunji1"))
        bunji2 = _pad_bunji(request.form.get("bunji2"))
        f = request.files.get("bank_file")
        if not (f and f.filename):
            old_token = (request.form.get("token") or "").strip()
            if old_token and _load_state(old_token):
                return redirect(
                    url_for(
                        "payments_import",
                        token=old_token,
                        bunji1=bunji1 or None,
                        bunji2=bunji2 or None,
                    )
                )
            return redirect(url_for("payments_import"))
        try:
            deposits, account_no = load_bank_deposits(f.filename, f.read())
        except Exception as e:
            flash(f"파일 읽기 실패: {e}", "err")
            return redirect(url_for("payments_import"))
        if not deposits:
            flash("입금 내역을 찾지 못했습니다 (파일 형식을 확인하세요).", "err")
            return redirect(url_for("payments_import"))

        auto_detected = False
        if not (bunji1 and bunji2):
            bunji1, bunji2 = detect_building(account_no)
            auto_detected = bool(bunji1 and bunji2)
            if not auto_detected:
                flash("건물을 자동으로 찾지 못했습니다. 직접 선택하세요.", "err")
        token = _save_state(
            bunji1,
            bunji2,
            deposits,
            auto_detected,
            filename=os.path.basename(f.filename or ""),
            account_no=account_no or "",
        )
        return redirect(url_for("payments_import", token=token))

    # GET: 최초 진입(빈 폼) 또는 방금 매칭한 결과 보기(?token=...),
    # 건물을 다시 골랐으면 ?token=...&bunji1=...&bunji2=...
    token = request.args.get("token") or ""
    state = _load_state(token) if token else None
    rows = None
    bunji1 = bunji2 = ""
    auto_detected = False
    uploaded_name = ""

    if state:
        req_b1 = _pad_bunji(request.args.get("bunji1"))
        req_b2 = _pad_bunji(request.args.get("bunji2"))
        if req_b1 and req_b2 and (req_b1, req_b2) != (state["bunji1"], state["bunji2"]):
            state = _update_state_building(token, req_b1, req_b2) or state
        bunji1, bunji2 = state["bunji1"], state["bunji2"]
        auto_detected = bool(state.get("auto_detected"))
        uploaded_name = (state.get("filename") or "").strip()
        rows = (
            _match_deposits(
                state["deposits"], bunji1, bunji2, state.get("account_no") or "",
            )
            if (bunji1 and bunji2)
            else []
        )

    return render_template(
        "payments_import.html",
        buildings=buildings,
        bunji1=bunji1,
        bunji2=bunji2,
        building_label=_building_label(bunji1, bunji2) if bunji1 and bunji2 else "",
        auto_detected=auto_detected,
        rows=rows,
        token=token,
        uploaded_name=uploaded_name,
        apply_err=(request.args.get("err") or "").strip(),
    )


@app.route("/payments/import/exclude", methods=["GET", "POST"])
@login_required
@require_write_access
def payments_import_exclude():
    """제외 목록 관리 — 글자(적요) + 선택 주소·계좌 범위.
    매칭 결과 화면(미매칭 행)에서 바로 제외할 때는 token을 같이 보내서
    등록 후 그 매칭 결과로 돌아가게 한다."""
    return_token = request.form.get("token") or ""

    def _back():
        if return_token:
            return redirect(url_for("payments_import", token=return_token))
        return redirect(url_for("payments_import_exclude"))

    if request.method == "POST" and request.form.get("action") == "exclude_add":
        add_exclude_keyword(
            request.form.get("keyword"),
            request.form.get("bunji1"),
            request.form.get("bunji2"),
            request.form.get("acct_no"),
        )
        return _back()
    if request.method == "POST" and request.form.get("action") == "exclude_edit":
        update_exclude_keyword(
            request.form.get("keyword_id"),
            request.form.get("keyword"),
            request.form.get("bunji1"),
            request.form.get("bunji2"),
            request.form.get("acct_no"),
        )
        return _back()
    if request.method == "POST" and request.form.get("action") == "exclude_del":
        try:
            delete_exclude_keyword(int(request.form.get("keyword_id") or 0))
        except ValueError:
            pass
        return _back()

    buildings, _rooms = _buildings_and_rooms()
    rows = list_exclude_keywords()
    q = (request.args.get("q") or "").strip()
    if q:
        ql = q.lower()
        def _blob(r):
            parts = [r.get("keyword") or "", r.get("juso") or "", r.get("acct_no") or "전체"]
            if r.get("bunji1"):
                parts.append(_fmt_bunji_pair(r.get("bunji1"), r.get("bunji2")))
            else:
                parts.append("전체")
            return " ".join(parts).lower()
        rows = [r for r in rows if ql in _blob(r)]
    pager = _make_pager(len(rows), _parse_page())
    page_rows = rows[pager["offset"] : pager["offset"] + pager["per_page"]]
    edit = None
    try:
        edit_id = int(request.args.get("edit_id") or 0)
    except ValueError:
        edit_id = 0
    if edit_id:
        edit = next((r for r in list_exclude_keywords() if int(r.get("id") or 0) == edit_id), None)
    return render_template(
        "payments_import_exclude.html",
        exclude_keywords=page_rows,
        buildings=buildings,
        edit=edit,
        pager=pager,
        q=q,
    )
